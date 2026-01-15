import pandas as pd
import json
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import logging
from typing import Dict
# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ReportGenerator:
    """Generate formatted Excel reports from LLM recommendations"""
    
    def __init__(self, recommendations: Dict, output_directory: str = './output'):
        self.recommendations = recommendations
        self.output_dir = Path(output_directory)
        self.output_dir.mkdir(exist_ok=True)
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    def create_campaign_adjustments_df(self) -> pd.DataFrame:
        """Convert campaign adjustments to DataFrame"""
        
        adjustments = self.recommendations['campaign_adjustments']
        
        # Flatten nested bid_adjustments
        flattened = []
        for item in adjustments:
            flat_item = {
                'campaign': item.get('campaign'),
                'default_bid_multiplier': item.get('default_bid_multiplier'),
                'bid_adj_top_of_search': item.get('bid_adjustments', {}).get('top_of_search'),
                'bid_adj_rest_of_search': item.get('bid_adjustments', {}).get('rest_of_search'),
                'bid_adj_product_pages': item.get('bid_adjustments', {}).get('product_pages'),
                'budget_action': item.get('budget_change', {}).get('action'),
                'budget_change_percent': item.get('budget_change', {}).get('percent'),
                'projected_daily_spend': item.get('projected_daily_spend_usd'),
                'projected_daily_sales': item.get('projected_daily_sales_usd'),
                'estimated_acos': item.get('estimated_acos_percent'),
                'estimated_roas': item.get('estimated_roas_multiple')
            }
            flattened.append(flat_item)
        
        df = pd.DataFrame(flattened)
        
        # Format currency columns
        currency_cols = ['projected_daily_spend', 'projected_daily_sales']
        for col in currency_cols:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: f"${x:.2f}" if pd.notna(x) else "")
        
        # Format percentage columns
        pct_cols = ['bid_adj_top_of_search', 'bid_adj_rest_of_search', 
                    'bid_adj_product_pages', 'budget_change_percent', 'estimated_acos']
        for col in pct_cols:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "")
        
        return df
    
    def create_keyword_recommendations_df(self) -> tuple:
        """Create DataFrames for add and negative keywords"""
        
        add_keywords = pd.DataFrame(
            self.recommendations['keyword_recommendations']['add_exact']
        )
        
        negative_keywords = pd.DataFrame(
            self.recommendations['keyword_recommendations']['negative']
        )
        
        # Format suggested_bid
        if 'suggested_bid' in add_keywords.columns:
            add_keywords['suggested_bid'] = add_keywords['suggested_bid'].apply(
                lambda x: f"${x:.2f}" if pd.notna(x) else ""
            )
        
        return add_keywords, negative_keywords
    
    def create_targeting_recommendations_df(self) -> pd.DataFrame:
        """Convert targeting recommendations to DataFrame"""
        
        df = pd.DataFrame(self.recommendations['targeting_recommendations'])
        
        # Format value column
        if 'value' in df.columns:
            df['value'] = df['value'].apply(
                lambda x: f"${x:.2f}" if pd.notna(x) else ""
            )
        
        return df
    
    def style_excel_worksheet(self, writer, sheet_name: str, df: pd.DataFrame):
        """Apply styling to Excel worksheet"""
        
        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for col_num, value in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Adjust column widths
        for idx, col in enumerate(df.columns, 1):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[get_column_letter(idx)].width = adjusted_width
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
    
    def generate_excel_report(self) -> Path:
        """Generate comprehensive Excel report"""
        
        output_file = self.output_dir / f'amazon_ads_recommendations_{self.timestamp}.xlsx'
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            
            # 1. Campaign Adjustments
            campaign_df = self.create_campaign_adjustments_df()
            campaign_df.to_excel(writer, sheet_name='Campaign Adjustments', index=False)
            self.style_excel_worksheet(writer, 'Campaign Adjustments', campaign_df)
            
            # 2. Keywords to Add
            add_keywords_df, negative_keywords_df = self.create_keyword_recommendations_df()
            add_keywords_df.to_excel(writer, sheet_name='Keywords - Add Exact', index=False)
            self.style_excel_worksheet(writer, 'Keywords - Add Exact', add_keywords_df)
            
            # 3. Negative Keywords
            negative_keywords_df.to_excel(writer, sheet_name='Keywords - Negative', index=False)
            self.style_excel_worksheet(writer, 'Keywords - Negative', negative_keywords_df)
            
            # 4. Targeting Adjustments
            targeting_df = self.create_targeting_recommendations_df()
            targeting_df.to_excel(writer, sheet_name='Targeting Adjustments', index=False)
            self.style_excel_worksheet(writer, 'Targeting Adjustments', targeting_df)
            
            # 5. Summary Sheet
            summary_df = pd.DataFrame({
                'Metric': [
                    'Total Campaigns Analyzed',
                    'Keywords to Add',
                    'Negative Keywords',
                    'Targeting Adjustments',
                    'Generated At'
                ],
                'Value': [
                    len(campaign_df),
                    len(add_keywords_df),
                    len(negative_keywords_df),
                    len(targeting_df),
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                ]
            })
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            self.style_excel_worksheet(writer, 'Summary', summary_df)
    def generate_json_report(self) -> Path:
        """Save raw JSON recommendations"""
        
        output_file = self.output_dir / f'recommendations_{self.timestamp}.json'
        
        with open(output_file, 'w') as f:
            json.dump(self.recommendations, f, indent=2)
        
        logger.info(f"✓ JSON report generated: {output_file}")
        return output_file

    def generate_summary_report(self, original_reports: Dict[str, pd.DataFrame]) -> Path:
        """Generate a text summary report"""
        
        output_file = self.output_dir / f'summary_{self.timestamp}.txt'
        
        with open(output_file, 'w') as f:
            f.write("="*80 + "\n")
            f.write("AMAZON ADS OPTIMIZATION SUMMARY\n")
            f.write("="*80 + "\n\n")
            f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            
            # Data Overview
            f.write("DATA OVERVIEW\n")
            f.write("-" * 40 + "\n")
            for report_type, df in original_reports.items():
                f.write(f"{report_type.title()}: {len(df)} rows\n")
            f.write("\n")
            
            # Recommendations Summary
            f.write("RECOMMENDATIONS SUMMARY\n")
            f.write("-" * 40 + "\n")
            f.write(f"Campaign Adjustments: {len(self.recommendations['campaign_adjustments'])}\n")
            f.write(f"Keywords to Add: {len(self.recommendations['keyword_recommendations']['add_exact'])}\n")
            f.write(f"Negative Keywords: {len(self.recommendations['keyword_recommendations']['negative'])}\n")
            f.write(f"Targeting Changes: {len(self.recommendations['targeting_recommendations'])}\n")
            f.write("\n")
            
            # Detailed Reasoning
            if 'Detailed reasoning' in self.recommendations:
                f.write("DETAILED REASONING\n")
                f.write("-" * 40 + "\n")
                f.write(f"{self.recommendations['Detailed reasoning']}\n\n")
            
            # Top Campaign Actions
            f.write("TOP CAMPAIGN ACTIONS\n")
            f.write("-" * 40 + "\n")
            for adj in self.recommendations['campaign_adjustments'][:5]:
                f.write(f"\nCampaign: {adj['campaign']}\n")
                budget = adj.get('budget_change', {})
                f.write(f"  Budget Action: {budget.get('action', 'none')}")
                percent_val = budget.get('percent')
                if percent_val is not None:
                    percent_str = f"{percent_val:.1f}%" if isinstance(percent_val, (int, float)) else f"{percent_val}%"
                    f.write(f" ({percent_str})")
                acos_val = adj.get('estimated_acos_percent', 'N/A')
                acos_str = f"{acos_val:.1f}%" if isinstance(acos_val, (int, float)) else f"{acos_val}%"
                f.write(f"\n  Est. ACOS: {acos_str}\n")
                
                roas_val = adj.get('estimated_roas_multiple', 'N/A')
                roas_str = f"{roas_val:.2f}x" if isinstance(roas_val, (int, float)) else f"{roas_val}"
                f.write(f"  Est. ROAS: {roas_str}\n")
        
        logger.info(f"✓ Summary report generated: {output_file}")
        return output_file